from openpyxl import load_workbook

def redact_excel(input_path, output_path):
    # Load workbook
    wb = load_workbook(input_path)
    
    # Iterate through all sheets
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                # Redact only text cells
                if isinstance(cell.value, str) and cell.value.strip() != "":
                    cell.value = "████████"

    # Save redacted workbook
    wb.save(output_path)
